"""Write research rows to Excel (addresses, locations, optional user_rating column)."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def rows_to_excel(rows: list[dict[str, Any]]) -> bytes:
    """
    Columns: activity_query, option_name, address, location, link, user_rating.
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active sheet")
    ws.title = "Research"

    headers = ["activity_query", "option_name", "address", "location", "link", "user_rating"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)

    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=row.get("activity_query", ""))
        ws.cell(row=r, column=2, value=row.get("option_name", ""))
        ws.cell(row=r, column=3, value=row.get("address", ""))
        ws.cell(row=r, column=4, value=row.get("location", ""))
        ws.cell(row=r, column=5, value=row.get("link", ""))
        ws.cell(row=r, column=6, value=row.get("user_rating", ""))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def rows_to_planning_json(rows: list[dict[str, Any]]) -> str:
    """JSON lines friendly for another app doing hour-by-hour planning."""
    import json

    return json.dumps(rows, indent=2, ensure_ascii=False)
