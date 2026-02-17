"""Export research results to an Excel workbook with columns for user rating."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# Schema: list of dicts with activity_query, option_name, address, location, link, (user_rating optional)
def research_results_to_excel(
    results: list[dict[str, Any]],
    *,
    include_user_rating_column: bool = True,
) -> bytes:
    """
    Write research results to an Excel file in memory.
    Columns: activity_query, option_name, address, location, link, user_rating (empty).
    Returns workbook bytes.
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    ws.title = "Research options"

    headers = ["activity_query", "option_name", "address", "location", "link"]
    if include_user_rating_column:
        headers.append("user_rating")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=row.get("activity_query") or "")
        ws.cell(row=row_idx, column=2, value=row.get("option_name") or "")
        ws.cell(row=row_idx, column=3, value=row.get("address") or "")
        ws.cell(row=row_idx, column=4, value=row.get("location") or "")
        ws.cell(row=row_idx, column=5, value=row.get("link") or "")
        if include_user_rating_column:
            ws.cell(row=row_idx, column=6, value=row.get("user_rating") or "")

    # Auto-fit would require openpyxl.utils; just set reasonable widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
